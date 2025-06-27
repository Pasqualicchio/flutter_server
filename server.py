from flask import Flask, request, redirect, url_for, render_template, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from PIL import Image
from pathlib import Path
from flask import send_from_directory
import os
import glob

# Configurazione dell'app Flask
app = Flask(__name__)
app.secret_key = 'supersecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024 * 1024  # 1 GB
db = SQLAlchemy(app)

# Percorsi base
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / 'uploads'
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
LOG_FILE = BASE_DIR / 'upload_log.txt'

# Configurazione email (modifica con i tuoi dati reali)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'tuo.email@gmail.com'
app.config['MAIL_PASSWORD'] = 'tua_app_password'
mail = Mail(app)

# Modello utente
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)

with app.app_context():
    db.create_all()

# Funzione utili
def safe_name(name):
    return secure_filename(name).replace(" ", "_").replace("-", "_")

# Decoratore login richiesto
from functools import wraps
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    allowed_extensions = {'jpg', 'jpeg', 'png', 'gif', 'mp4', 'webm', 'ogg'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

# Aggiorna video in cartelle
# Funzione per salvare i file in cartelle strutturate e aggiungere l'anteprima
def save_file_in_folders(macro, commessa, matricola, item, file):
    today = datetime.now().strftime('%Y%m%d')  # Data corrente per il nome file

    # Creiamo la struttura delle cartelle
    macro_folder = os.path.join(app.config['UPLOAD_FOLDER'], safe_name(macro))  # Cartella per la macro categoria
    commessa_folder = os.path.join(macro_folder, f"V{commessa}")  # Cartella per la commessa
    matricola_folder = os.path.join(commessa_folder, f"Matr.{matricola}")  # Cartella per la matricola
    item_folder = os.path.join(matricola_folder, f"It.{item}")  # Cartella per l'item

    # Creiamo le cartelle se non esistono
    os.makedirs(item_folder, exist_ok=True)

    # Nome del file sicuro
    ext = os.path.splitext(file.filename)[1].lower()
    filename = f"{today}_V{safe_name(commessa)}_Matr.{safe_name(matricola)}_It.{safe_name(item)}{ext}"
    file_path = os.path.join(item_folder, filename)  # Percorso completo per il salvataggio

    # Salviamo il file
    file.save(file_path)

    return file_path, filename

# Funzione per aggiornare l'Excel con anteprima per immagini e video
def update_excel_file(macro_folder, macro_name, log_data, file_path):
    excel_path = Path(macro_folder) / f"log_{safe_name(macro_name)}.xlsx"
    headers = ["Timestamp", "Macro", "Linea", "Commessa", "Matricola", "Item", "Filename", "Preview"]

    # Se il file Excel non esiste, creiamo un nuovo file con l'intestazione
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    # Aggiungi i dati di log
    row_idx = ws.max_row + 1
    for col_idx, value in enumerate(log_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Verifica se il file è un'immagine
    file_path = Path(file_path)
    ext = file_path.suffix.lower()

    if ext in ['.jpg', '.jpeg', '.png', '.gif']:  # Se il file è un'immagine
        try:
            img = XLImage(str(file_path))  # Aggiungi l'immagine
            img.width = 100
            img.height = 70
            ws.add_image(img, f"H{row_idx}")  # Posiziona l'immagine nella colonna H
            ws.row_dimensions[row_idx].height = 55  # Imposta l'altezza della riga
        except Exception as e:
            print(f"[Errore immagine Excel] {e}")

    # Se il file è un video, aggiungi un'icona per il video
    elif ext in ['.mp4', '.webm', '.ogg', '.avi', '.mov']:
        try:
            print(f"[Debug] Aggiunta dell'icona video per: {file_path.name}")  # Debug: mostra il nome del file video

            video_icon_path = Path(BASE_DIR, 'static', 'video_icon.png')  # Percorso dell'icona video
            print(f"[Debug] Verifica dell'icona video in: {video_icon_path}")

            if video_icon_path.exists():
                video_icon = XLImage(str(video_icon_path))  # Aggiungi l'icona del video
                video_icon.width = 100
                video_icon.height = 70
                ws.add_image(video_icon, f"H{row_idx}")  # Posiziona l'icona nella colonna H
                ws.row_dimensions[row_idx].height = 55  # Imposta l'altezza della riga
            else:
                print(f"[Errore] L'icona del video non esiste nel percorso: {video_icon_path}")

        except Exception as e:
            print(f"[Errore icona video Excel] {e}")

    # Aggiorna la larghezza delle colonne
    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=row, column=col).value or '')) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 15)

    # Salva il file Excel
    wb.save(excel_path)




# Funzione aggiornata per ottenere informazioni sui file
def get_files_info(folder_path, allowed_exts=None):
    files = []
    # Verifica se la cartella esiste
    if not os.path.exists(folder_path):
        return []

    for root, _, filenames in os.walk(folder_path):
        for filename in filenames:
            ext = os.path.splitext(filename)[1].lower()

            # Se sono definite estensioni per cui permettere il filtro
            if allowed_exts and ext not in allowed_exts:
                continue

            full_path = os.path.join(root, filename)
            stat = os.stat(full_path)

            # Aggiungi il tipo di file
            file_type = 'other'
            if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                file_type = 'image'
            elif ext in ['.mp4', '.avi', '.mov', '.mkv', '.webm', '.ogg']:  # Aggiungi più formati video se necessario
                file_type = 'video'
            elif ext in ['.xlsx', '.xls']:
                file_type = 'excel'
            elif ext == '.pdf':
                file_type = 'pdf'
            elif ext in ['.mp3', '.wav', '.aac']:
                file_type = 'audio'

            files.append({
                'path': full_path.replace("\\", "/"),  # per compatibilità Windows
                'name': filename,
                'type': file_type,
                'created': datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S"),
                'size_kb': round(stat.st_size / 1024, 2)  # Dimensione in KB
            })

    return files


def paginate(items, page=1, per_page=20):
    total_items = len(items)
    total_pages = (total_items + per_page - 1) // per_page
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], total_pages

@app.route('/browse-ui')
def browse_ui():
    images_info = get_files_info(UPLOAD_FOLDER)
    excels = get_files_info(UPLOAD_FOLDER, allowed_exts=['.xlsx'])
    page = 1
    per_page = 20
    paginated_images, total_pages = paginate(images_info, page=page, per_page=per_page)
    return render_template('browse.html', images=paginated_images, excels=excels, page=page, total_pages=total_pages)

# Pagina principale
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('upload_advanced_ui'))
    return redirect(url_for('login'))

# Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            return redirect(url_for('upload_advanced_ui'))
        return render_template('login.html', error="❌ Credenziali errate")
    return render_template('login.html')

# Registrazione
@app.route('/register', methods=['GET', 'POST'])
def new_register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not username or not password:
            return render_template('register.html', error="Compila tutti i campi")
        if password != confirm_password:
            return render_template('register.html', error="Le password non corrispondono.")

        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            return render_template('register.html', error="L'utente esiste già.")

        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('register.html')

# Logout
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# Pagina upload
@app.route('/upload-advanced-ui')
@login_required
def upload_advanced_ui():
    return render_template('uploadadvancedui.html')

# Caricamento immagini/video
@app.route('/upload', methods=['POST'])
@login_required
def upload():
    try:
        # 1. Dati dal form (strip per sicurezza)
        macro     = request.form.get('macro', '').strip()
        linea     = request.form.get('linea', '').strip()
        commessa  = request.form.get('commessa', '').strip()
        matricola = request.form.get('matricola', '').strip()
        item      = request.form.get('item', '').strip()
        images    = request.files.getlist('image')
        videos    = request.files.getlist('video')

        # 2. Validazione
        if not all([macro, commessa, matricola, item]) or (not images and not videos):
            return jsonify({'status': 'error', 'message': 'Dati mancanti o nessun file'}), 400

        # 3. Cartella principale per macro
        macro_folder = os.path.join(UPLOAD_FOLDER, safe_name(macro))
        os.makedirs(macro_folder, exist_ok=True)

        # 4. Cartella finale con prefissi (senza safe_name, per mantenere i punti)
        photo_folder = os.path.join(
            macro_folder,
            f"V{commessa}",
            f"Matr.{matricola}",
            f"It.{item}"
        )
        # Debug: stampa il percorso per verificare
        print(f"📁 Creazione cartella: {photo_folder}")
        os.makedirs(photo_folder, exist_ok=True)

        uploaded_files = []
        today = datetime.now().strftime('%Y%m%d')

        # 5. Nome file base (sicuro)
        def base_filename():
            return f"{today}_V{safe_name(commessa)}_Matr.{safe_name(matricola)}_It.{safe_name(item)}"

        # 6. Contatore anti-sovrascrittura
        def build_unique_filename(base_name, ext, folder):
            counter = 1
            filename = f"{base_name}{ext}"
            while os.path.exists(os.path.join(folder, filename)):
                filename = f"{base_name}_{counter}{ext}"
                counter += 1
            return filename

        # 7. Salvataggio immagini
        for image in images:
            ext = os.path.splitext(image.filename)[1].lower()
            base = base_filename()
            filename = build_unique_filename(base, ext, photo_folder)
            image_path = os.path.join(photo_folder, filename)
            image.save(image_path)

            # Aggiorna Excel con i dati del file
            log_data = [datetime.now().isoformat(), macro, linea, commessa, matricola, item, filename]
            update_excel_file(macro_folder, macro, log_data, image_path)
            uploaded_files.append(filename)

        # 8. Salvataggio video
        for video in videos:
            ext = os.path.splitext(video.filename)[1].lower()
            base = base_filename()
            filename = build_unique_filename(base, ext, photo_folder)
            video_path = os.path.join(photo_folder, filename)
            video.save(video_path)
            
            # Aggiorna Excel con i dati del file
            log_data = [datetime.now().isoformat(), macro, linea, commessa, matricola, item, filename]
            update_excel_file(macro_folder, macro, log_data, video_path)
            uploaded_files.append(filename)

        return jsonify({'status': 'success', 'filenames': uploaded_files}), 200

    except Exception as e:
        # Cattura eventuali errori e loggali per il debug
        print(f"[Errore upload] {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/preview-all')
@login_required
def preview_all_images():
    all_files = get_files_info(UPLOAD_FOLDER)
    media_files = [f for f in all_files if f['type'] == 'image']
    return render_template('preview_images.html', files=media_files, upload_folder=str(UPLOAD_FOLDER))

@app.route('/preview-images/<path:excel_filename>')
@login_required
def preview_excel_images(excel_filename):
    full_excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)

    if not os.path.exists(full_excel_path):
        return f"File {full_excel_path} non trovato", 404

    image_folder = os.path.dirname(full_excel_path)

    image_paths = []

    for root, dirs, files in os.walk(image_folder):
        for file in files:
            if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                relative_path = os.path.relpath(os.path.join(root, file), UPLOAD_FOLDER).replace("\\", "/")
                image_paths.append(relative_path)

    # Aggiungi debug
    print(f"Percorsi immagini trovati: {image_paths}")

    # Se non ci sono immagini
    if not image_paths:
        return "Nessuna immagine trovata."

    grouped_images = {}
    for path in image_paths:
        parts = path.split("/")
        commessa = parts[1]  # Es. 'V3'
        matricola = parts[2]  # Es. 'Matr.3'
        item = parts[3]  # Es. 'It.3'
        
        # Aggiungi debug per ogni gruppo
        print(f"Gruppo: {commessa}, {matricola}, {item}")
        
        if (commessa, matricola, item) not in grouped_images:
            grouped_images[(commessa, matricola, item)] = []
        
        grouped_images[(commessa, matricola, item)].append(path)

    # Aggiungi un debug per vedere i gruppi di immagini
    print(f"Immagini raggruppate: {grouped_images}")

    # Calcolare la paginazione
    per_page = 20
    total_images = len(image_paths)
    total_pages = (total_images + per_page - 1) // per_page  # arrotonda per eccesso

    page = request.args.get('page', 1, type=int)  # Imposta la pagina corrente tramite il parametro URL
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_images = image_paths[start_idx:end_idx]

    return render_template("preview_images.html", grouped_images=grouped_images, page=page, total_pages=total_pages, files=paginated_images)

@app.route('/download/<path:path>')
def download_file(path):
    return send_from_directory(str(UPLOAD_FOLDER), path, as_attachment=True)

@app.route('/serve-image/<path:filename>')
@login_required
def serve_image(filename):
    return send_from_directory(str(UPLOAD_FOLDER), filename)

@app.route('/dropanddrag')
@login_required
def drop_and_drag_ui():
    return render_template('dropanddrag.html')


@app.route('/upload-drag', methods=['POST'])
@login_required
def upload_drag():
    try:
        # Log dei file ricevuti per il debug
        print("📸 Ricevuti file:", request.files.getlist('images[]'))
        
        # Estrazione dei dati dal form
        macro     = request.form.get('macro', '').strip()
        linea     = request.form.get('linea', '').strip()
        commessa  = request.form.get('commessa', '').strip()
        matricola = request.form.get('matricola', '').strip()
        item      = request.form.get('item', '').strip()
        files     = request.files.getlist('images[]')  # Può includere sia immagini che video
        
        # Validazione dei dati
        if not all([macro, commessa, matricola, item]) or not files:
            return jsonify({'status': 'error', 'message': 'Campi obbligatori mancanti o nessun file selezionato'}), 400

        # Creazione delle cartelle per salvare i file
        macro_folder = os.path.join(app.config['UPLOAD_FOLDER'], safe_name(macro))
        photo_folder = os.path.join(macro_folder, f"V{commessa}", f"Matr.{matricola}", f"It.{item}")
        os.makedirs(photo_folder, exist_ok=True)

        today = datetime.now().strftime('%Y%m%d')

        def base_filename():
            return f"{today}_V{safe_name(commessa)}_Matr.{safe_name(matricola)}_It.{safe_name(item)}"

        def build_unique_filename(base_name, ext, folder):
            """Crea un nome unico per il file, evitando sovrascrittura"""
            counter = 1
            filename = f"{base_name}{ext}"
            while os.path.exists(os.path.join(folder, filename)):
                filename = f"{base_name}_{counter}{ext}"
                counter += 1
            return filename

        uploaded_files = []  # Lista dei file caricati con successo

        # Elaborazione dei file ricevuti (immagini e video)
        for f in files:
            if f:
                ext = os.path.splitext(f.filename)[1].lower()
                if not allowed_file(f.filename):  # Verifica che il file sia permesso
                    return jsonify({'status': 'error', 'message': f'File {f.filename} non valido'}), 400

                base = base_filename()
                filename = build_unique_filename(base, ext, photo_folder)
                filepath = os.path.join(photo_folder, filename)
                
                # Salvataggio del file
                f.save(filepath)

                # Aggiornamento del file Excel
                update_excel_file(
                    macro_folder,
                    macro,
                    [datetime.now().isoformat(), macro, linea, commessa, matricola, item, filename],
                    filepath
                )

                uploaded_files.append(filename)  # Aggiungi il nome del file alla lista

            else:
                # Se il file è vuoto, restituisci un errore
                return jsonify({'status': 'error', 'message': 'File vuoto ricevuto'}), 400

        # Risposta di successo con i file caricati
        return jsonify({'status': 'success', 'files': uploaded_files}), 200

    except Exception as e:
        print(f"[Errore upload-drag] {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# Avvio del server
if __name__ == "__main__":
    app.run(debug=True, port=5005, use_reloader=False)




