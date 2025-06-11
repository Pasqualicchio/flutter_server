from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import os, time

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# MODELLO UTENTE
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)

with app.app_context():
    db.create_all()


# 🔽 FUNZIONE: Recupera file nella cartella 'uploads', ordinati per data
def get_all_files():
    folder = "uploads/"
    files = []

    if not os.path.exists(folder):
        os.makedirs(folder)

    for fname in os.listdir(folder):
        path = os.path.join(folder, fname)
        if os.path.isfile(path):
            created = os.path.getmtime(path)
            ext = fname.lower()
            if ext.endswith(('.png', '.jpg', '.jpeg', '.gif')):
                ftype = 'image'
            elif ext.endswith(('.mp4', '.mov', '.avi')):
                ftype = 'video'
            elif ext.endswith('.pdf'):
                ftype = 'pdf'
            elif ext.endswith(('.mp3', '.wav')):
                ftype = 'audio'
            elif ext.endswith(('.xls', '.xlsx')):
                ftype = 'excel'
            else:
                ftype = 'other'

            files.append({
                'path': fname,
                'type': ftype,
                'created': created,
                'size_kb': os.path.getsize(path) // 1024
            })

    files.sort(key=lambda x: x['created'], reverse=True)
    return files

# HOME
@app.route('/')
def home():
    return render_template('home.html')

# LOGIN
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            return redirect(url_for('upload_advanced_ui'))
        return "❌ Credenziali errate", 401
    return render_template('login.html')

# REGISTRAZIONE
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm = request.form['confirm_password']
        if password != confirm:
            return "❌ Le password non coincidono", 400
        if User.query.filter_by(username=username).first():
            return "❌ Username già registrato", 400
        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('register.html')

# UPLOAD (GET + POST)
@app.route('/upload', methods=['GET', 'POST'])
def upload_advanced_ui():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))

    if request.method == 'POST':
        macro = request.form.get('macro')
        commessa = request.form.get('commessa')
        matricola = request.form.get('matricola')
        item = request.form.get('item')
        note = request.form.get('note', '').strip().replace(" ", "-")  # opzionale, sanificato

        if not all([macro, commessa, matricola, item]):
            return jsonify({"status": "error", "message": "Compila tutti i campi"}), 400

        files = request.files.getlist('images[]')
        if not files or files == [None] or files == []:
            single = request.files.get('file')
            if single and single.filename:
                files = [single]

        if not files or all(f.filename == '' for f in files):
            return jsonify({"status": "error", "message": "Nessun file valido selezionato"}), 400

        upload_folder = "uploads"
        os.makedirs(upload_folder, exist_ok=True)

        salvati = []
        single_upload = len(files) == 1

        for i, file in enumerate(files, start=1):
            if file and file.filename:
                ext = file.filename.rsplit('.', 1)[-1].lower()
                note_suffix = f"_{note}" if note else ""

                if single_upload:
                    filename = f"{macro}_{commessa}_{matricola}_{item}{note_suffix}.{ext}"
                else:
                    filename = f"{macro}_{commessa}_{matricola}_{item}_n{i}{note_suffix}.{ext}"

                path = os.path.join(upload_folder, filename)
                file.save(path)
                salvati.append(filename)

        return jsonify({
            "status": "success",
            "message": f"{len(salvati)} file caricati con successo!",
            "files": salvati
        })

    return render_template('upload_advanced_ui.html', user=user)


# BROWSE
@app.route('/browse')
def browse():
    # Recupera tutti i file nella cartella 'uploads'
    files = get_all_files()
    # Filtro per immagini e video
    images = [f for f in files if f['type'] in ['image', 'video']]
    # Filtro per file Excel
    excels = [f for f in files if f['type'] == 'excel']

    # Recupera la pagina attuale, se non è presente si usa la pagina 1
    page = int(request.args.get('page', 1))
    
    # Numero di file per pagina
    per_page = 12
    # Calcola il numero totale di pagine
    total_pages = (len(images) + per_page - 1) // per_page

    # Se la pagina è maggiore di quelle disponibili, la imposta sull'ultima pagina
    if page > total_pages:
        page = total_pages

    # Calcola l'intervallo di file da visualizzare per la pagina corrente
    start = (page - 1) * per_page
    end = start + per_page
    paginated_files = images[start:end]

    # Rendi la pagina con i file paginati
    return render_template('browse.html', images=paginated_files, excels=excels, page=page, total_pages=total_pages)


# DOWNLOAD GLOBALE EXCEL
@app.route('/download-global-excel')
def download_global_excel():
    os.makedirs("exports", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Log Completo"

    headers = [
        "Macro", "Commessa", "Matricola", "Item", 
        "Nome File", "Tipo", "Data", "Dimensione (KB)", "Anteprima"
    ]
    ws.append(headers)

    # Stili
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Applica stile all’intestazione
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    # Filtro automatico
    ws.auto_filter.ref = f"A1:I1"

    # Larghezza colonne suggerita
    col_widths = [12, 15, 15, 12, 35, 10, 20, 15, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    row_index = 2
    for f in get_all_files():
        filename = f['path']
        parts = filename.rsplit("_", 4)

        macro = commessa = matricola = item = "N/D"
        if len(parts) >= 4:
            macro = parts[0]
            commessa = parts[1]
            matricola = parts[2]
            item_raw = parts[3]
            item = item_raw.split("_n")[0]

        data_row = [
            macro,
            commessa,
            matricola,
            item,
            filename,
            f['type'],
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(f['created'])),
            f['size_kb'],
            ""  # Placeholder immagine
        ]

        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_index, column=col_idx, value=value if col_idx != 9 else None)
            cell.border = thin_border
            cell.alignment = center_align
            if row_index % 2 == 0:
                cell.fill = zebra_fill

        # Inserisci immagine nella colonna I
        if f['type'] == 'image':
            img_path = os.path.join('uploads', f['path'])
            try:
                with PILImage.open(img_path) as pil_img:
                    pil_img.thumbnail((90, 90))
                    thumb_path = os.path.join("exports", f"thumb_{f['path']}")
                    pil_img.save(thumb_path)

                excel_img = ExcelImage(thumb_path)
                excel_img.width, excel_img.height = 90, 90
                ws.row_dimensions[row_index].height = 80
                ws.add_image(excel_img, f'I{row_index}')
            except Exception as e:
                print(f"Errore immagine {img_path}: {e}")

        row_index += 1

    output_path = "exports/global_log.xlsx"
    wb.save(output_path)
    return send_file(output_path, as_attachment=True)

# DOWNLOAD FILE SINGOLO
@app.route('/download/<path:filename>')
def download_file(filename):
    path = os.path.join("uploads", filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "❌ File non trovato", 404

# LOGOUT
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)