from flask import Flask, request, redirect, url_for, render_template, session, jsonify, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from functools import wraps
import sqlite3
import os
import math
import time
from pathlib import Path
from werkzeug.utils import secure_filename
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy

# Inizializzazione app Flask e SQLAlchemy
app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Classe User per gestire l'autenticazione
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)

# Crea il database se non esiste
with app.app_context():
    db.create_all()

# Configurazione per l'upload dei file e per il server email
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / 'uploads'
UPLOAD_FOLDER.mkdir(exist_ok=True)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'tuo.email@gmail.com'  # Cambia con il tuo indirizzo
app.config['MAIL_PASSWORD'] = 'tua_app_password'      # Cambia con la tua password/app password

mail = Mail(app)

# Funzione per la paginazione dei risultati
def paginate(items, page=1, per_page=20):
    total = len(items)
    total_pages = math.ceil(total / per_page)
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], total_pages

# Funzione per sanificare il nome dei file
def safe_name(name):
    return secure_filename(name).replace(" ", "_").replace("-", "_")

# Decoratore per proteggere le route che richiedono l'autenticazione
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))  # Reindirizza alla pagina di login se non autenticato
        return f(*args, **kwargs)
    return decorated_function

# Funzione per ottenere informazioni sui file nella cartella
def get_files_info(folder, allowed_exts=None):
    result = []
    for root, _, files in os.walk(folder):
        for file in files:
            if allowed_exts and not file.lower().endswith(tuple(allowed_exts)):
                continue
            full_path = Path(root) / file
            rel_path = full_path.relative_to(folder)
            stat = full_path.stat()
            ext = file.lower().split('.')[-1]
            file_type = 'image' if ext in ['jpg', 'jpeg', 'png'] else \
                        'video' if ext in ['mp4', 'avi'] else \
                        'pdf' if ext == 'pdf' else \
                        'audio' if ext in ['mp3', 'wav'] else 'other'
            result.append({
                'name': file,
                'path': str(rel_path).replace("\\", "/"),
                'type': file_type,
                'size_kb': round(stat.st_size / 1024, 1),
                'created': datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M"),
                'modified': datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")
            })
    return result

def wait_for_unlock(path, timeout=5):
    lock_path = Path(str(path) + ".lock")
    start = time.time()
    while lock_path.exists():
        if time.time() - start > timeout:
            raise TimeoutError("Timeout: file ancora bloccato")
        time.sleep(0.1)
    lock_path.touch()  # crea il file lock

def release_lock(path):
    lock_path = Path(str(path) + ".lock")
    if lock_path.exists():
        lock_path.unlink()

# Funzione per ottenere tutti i file dalla cartella di upload
def get_all_files():
    folder = UPLOAD_FOLDER  # Cartella dove vengono salvati i file
    files = []
    for root, dirs, filenames in os.walk(folder):
        for filename in filenames:
            file_path = os.path.join(root, filename)
            file_info = {
                'path': file_path,
                'name': filename,
                'type': 'other'  # Puoi migliorare la logica per determinare il tipo di file (immagine, video, ecc.)
            }
            files.append(file_info)
    return files

def get_last_modified_time():
    global_log_path = BASE_DIR / 'global_log.xlsx'
    if global_log_path.exists():
        return datetime.fromtimestamp(global_log_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    return None

# Route di login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            return redirect(url_for('upload_advanced_ui_func'))
        return "❌ Credenziali errate", 401
    return render_template('login.html')

# Route di registrazione
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm = request.form['confirm_password']
        if password != confirm:
            return "❌ Le password non coincidono", 400
        if User.query.filter_by(username=username).first():
            return "❌ Username già registrato", 400
        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('register.html')

# Pagina principale
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('browse'))
    return redirect(url_for('login'))

# Funzione di logout
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# Funzione browse (per visualizzare i file)
@app.route('/browse')
def browse():
    files = get_all_files()
    images = [f for f in files if f['type'] in ['image', 'video']]
    excels = [f for f in files if f['type'] == 'excel']

    last_modified_time = get_last_modified_time()

    # Paginazione immagini
    page = int(request.args.get('page', 1))
    per_page = 12
    total_pages = (len(images) + per_page - 1) // per_page or 1  # Evita la divisione per zero

    page = min(page, total_pages)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_files = images[start:end]

    return render_template(
        'browse.html',
        images=paginated_files,
        excels=excels,
        page=page,
        total_pages=total_pages,
        last_modified_time=last_modified_time
    )

# Funzione di upload
@app.route('/upload', methods=['POST'])
def upload_func():
    files = request.files.getlist('images[]')  # Prendi la lista di file

    if not files or all(f.filename == '' for f in files):  # Verifica se ci sono file validi
        return jsonify({"status": "error", "message": "Nessun file valido selezionato"}), 400

    saved_files = []  # Lista per i file salvati

    for file in files:
        if file and file.filename:
            save_path = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))  # Salva con nome sicuro
            file.save(save_path)
            saved_files.append(file.filename)  # Aggiungi il nome del file alla lista

    return jsonify({
        "status": "success",
        "message": f"{len(saved_files)} file caricati con successo!",
        "files": saved_files
    }), 200

# Funzione per la pagina di upload avanzato
@app.route('/upload-advanced-ui', methods=['GET', 'POST'])
def upload_advanced_ui_func():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))

    if request.method == 'POST':
        pass

    return render_template('uploadadvancedui.html', user=user)

@app.route('/preview-images/<path:excel_filename>')
@login_required
def preview_excel_images(excel_filename):
    full_excel_path = UPLOAD_FOLDER / excel_filename

    if not full_excel_path.exists():
        return f"File {full_excel_path} non trovato", 404

    image_folder = full_excel_path.parent
    image_paths = []

    for root, _, files in os.walk(image_folder):
        for file in files:
            if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                full_path = os.path.join(root, file)
                relative_path = os.path.relpath(full_path, start=UPLOAD_FOLDER)
                image_paths.append(relative_path.replace("\\", "/"))  # importante per Windows

    return render_template("preview_images.html", files=image_paths)

# Avvio del server Flask
if __name__ == "__main__":
    app.run(debug=True, port=5006)
